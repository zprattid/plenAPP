##Librerías necesarias para extraer datos del PDF
from tika import parser
import re

from configuraciones import *

from openpyxl import load_workbook
import datetime

class Reader:
	def PDFtoTXT(pdf):
		pass
	def PDFtoINCIDENCIA(txt):
		print("> Leyendo Informe en PDF")
		file_data = parser.from_file(txt)
		text = file_data['content']
		txt = text.split("EVENTO GENERADO POR OPERADOR")
		incARRAY = []
		for item in txt[1:]:
			iName = None
			iHora = None
			iFecha = None
			iComments = []
			lines = item.split("\n")
			for line in lines:
				if line == "":
					pass
				else:
					#Obtener hora y nombre
					if iName == None:
						hora = re.search(r'\d\d:\d\d:\d\d OIL',line)
						try:
							iHora = hora.group()[:-7]
							iName = line.split(" - ")[-1][:-13]
							#print(iHora)
						except AttributeError:
							pass
					else:
						iComments.append(line)
					if iFecha == None:	
						fecha = re.search(r'\d\d/\d\d/\d\d\d\d',line)
						try:
							iFecha = fecha.group()[:-2]
							#print(line)
						except AttributeError:
							pass
					'''ctrl = re.search(r'\d\d/\d\d/\d\d\d\d\n',line)
					try:
						c = ctrl.group()[:-2]
						print(line)
					except AttributeError:
						pass'''
			#print(">>>>>"+str(iName)+"-"+str(iHora)+" "+str(iFecha))
			#print(iComments)
			incARRAY.append(Incidencia(iName, iHora, iFecha, iComments))
		return incARRAY
	def XLStoINCIDENCIA(xls):
		print("> Leyendo Excel incidencias")
		iName = None
		iHora = None
		iFecha = None
		iComments = None
		arrayINC = []
		wb = load_workbook(filename = xls)
		#print(wb.worksheets)
		ranges = 3
		for i in range(ranges):
			ws = wb.worksheets[i]
			print(">> Leyendo hoja: "+str(i))
			###########COLUMNA HORA#######################
			col = ws["C"]
			#print(len(col))
			for x in range(len(col)):
				#print(str(type(col[x].value))+"||"+str(col[x].value))
				if type(col[x].value) == str:
					hora = re.search(r'\d\d:\d\d',col[x].value)
					try:
						if hora.group():
							iHora = hora.group()
					except:
						'''print(str(type(col[x].value))+"||"+str(col[x].value))
						input()'''
						pass
				elif type(col[x].value) == datetime.time:
					hour = col[x].value.hour
					minute = col[x].value.minute
					if hour < 10:
						hour = "0"+str(hour)
					else:
						hour = str(hour)
					if minute < 10:
						minute = "0"+str(minute)
					else:
						minute = str(minute)
					iHora = hour+":"+minute
			###########COLUMNA FECHA#######################
				if type(ws["B"][x].value) == str:
					fecha = re.search(r'\d\d/\d\d/\d\d',ws["B"][x].value)
					try:
						if fecha.group():
							iFecha = fecha.group()
					except:
						'''print(str(type(col[x].value))+"||"+str(col[x].value))
						input()'''
						pass
				elif type(ws["B"][x].value) == datetime.datetime:
					day = 0
					month = 0
					if ws["B"][x].value.day < 10:
						day = "0"+str(ws["B"][x].value.day)
					else:
						day = str(ws["B"][x].value.day)
					if ws["B"][x].value.month < 10:
						month = "0"+str(ws["B"][x].value.month)
					else:
						month = str(ws["B"][x].value.month)
					iFecha = day+"/"+month+"/"+str(ws["B"][x].value.year)[2:]
			##########COLUMNA NOMBRE###########
				try:
					iName = ws["A"][x].value
				except KeyError:
					pass
				if iName == "ESTACION":
					pass
				else:
					arrayINC.append(Incidencia(iName,iHora,iFecha,iComments))
		return arrayINC 
	def FILEtoOBJ(adjunto):
		if adjunto[-4:] == ".pdf":
			return Reader.PDFtoINCIDENCIA(adjunto)
		elif adjunto[-5:] == ".xlsx":
			return Reader.XLStoINCIDENCIA(adjunto)

class Incidencia:
	def _checkExcluido(self):
		#Busca cadenas de texto en la incidencia para activar el flag self.excluido
		if self.comments is not None:
			for line in self.comments:
				if " ANULA " in line:
					self.excluido = True
				elif " ANULADO" in line:
					self.excluido = True
				elif " POR ERROR " in line:
					self.excluido = True
				elif " COMPROBACION " in line:
					self.excluido = True
				elif " COMPROBACIÓN " in line:
					self.excluido = True
	def __init__(self, nombre, hora, fecha, comments):
		self.nombre = nombre
		self.hora = hora
		self.fecha = fecha
		self.comments = comments
		self.excluido = False
		self.inExcel = False
		self._checkExcluido()

class Contador:
	def __init__(self, informe, excel, copia):
		self.objArray = Reader.FILEtoOBJ(informe) #incidencias cargadas desde informe
		self.excelArray = Reader.FILEtoOBJ(excel) #incidencias cargadas desde excel
		self.incidenciasInforme = 0 #cantidad de incidencias en el informe
		self.excluidos = 0 #incidencias excluidas del excel por su tipo
		self.incidenciasADD = 0 #incidencias no excluidas no apuntadas en el excel
		self.incidenciasERROR = 0 #incidencias en el excel pero que no coinciden con ninguna del informe
		self.incidenciasExcel = 0 #cuenta de las incidencias totales apuntadas en el excel
		self.incidenciasCalculadas = 0 #incidenciasInforme-excluidos
	def calcIncidencias(self):
		#Calcula las estadisticas.
		self.incidenciasInforme = len(self.objArray)
		self.incidenciasExcel = len(self.excelArray)
		for i in range(len(self.objArray)):
			if self.objArray[i].excluido == False and self.objArray[i].inExcel == False:
				self.incidenciasADD = self.incidenciasADD+1
			if self.objArray[i].excluido == True:
				self.excluidos = self.excluidos+1
		self.incidenciasCalculadas = self.incidenciasInforme-self.excluidos
	def compareExcel(self):
		#Itera sobre excelArray y lo compara con cada elemento de objArray
		#Si no está, se pone en amarillo
		#Si existe, se pone en verde y se activa la flag "inExcel" de objArray
		for inc in self.excelArray:
			for item in self.objArray:
				if item.excluido == False:
					#print(inc.fecha+" - "+item.fecha)
					if inc.fecha == item.fecha:
						if inc.hora == item.hora:
						#poner en verde
							item.inExcel = True
							inc.inExcel = True
						else:
							pass
							#print("> Check Fecha: "+str(inc.fecha)+" - "+str(item.fecha))
							#print("> Check Hora: "+str(inc.hora)+" - "+str(item.hora))
					else:
						#poner en amarillo. Suma 1 a incidenciasERROR
						#self.incidenciasERROR = self.incidenciasERROR+1
						pass
						##print(inc.hora+" - "+item.hora)
				else:
					pass
					
		pass
	def fillCopy(self):
		pass
	def addToExcel(self):
		#Añade las incidencias que no estan excluidas, ni en el excel
		for item in self.objArray:
			if item["excluido"] == False and item["inExcel"] == False:
				#add to excel
				#añadir en rojo
				pass


a = Contador("INFORME18.pdf","INCIDENCIAS PLENOIL.xlsx","copy.xlsx")
a.calcIncidencias()
a.compareExcel()
print("Total: "+str(a.incidenciasInforme))
print("Excluidos: "+str(a.excluidos))
print("Incidencias calculadas: "+str(a.incidenciasCalculadas))
print("Incidencias excel: "+str(a.incidenciasExcel))
print("ERRORES: "+str(a.incidenciasERROR))
count = 0
for i in a.objArray:
	if i.inExcel == True:
		count = count +1
print("Coincidentes: "+str(count))
count = 0
for i in a.objArray:
	if i.inExcel == False and i.excluido == False:
		count = count +1
		print(">>> "+i.nombre+" - "+i.fecha+" - "+i.hora)
		'''print(i.comments)'''
print("No coincidentes: "+str(count))
'''for i in a.excelArray:
	print(">>> "+i.nombre)
	print(">>> "+i.fecha+" - "+i.hora)
	print(i.comments)'''
