# -*- coding: utf-8 -*-

import re
from openpyxl import load_workbook
import os
import datetime

class plenFLUX:
	def __init__(self):
		self.excelPATH = "\\\\192.168.102.5\\t. de noche\\EXCEL PLENOIL\\"
		excelNAME = "\\\\192.168.102.5\\t. de noche\\EXCEL PLENOIL\\INCIDENCIAS PLENOIL.xlsx"
		self.excelNAME = "Flujo de clientes.xlsx"
		self.paths = []
		files = os.listdir(self.excelPATH)
		for i in files:
			if "INCIDENCIAS PLENOIL" in i:
				self.paths.append(i)
		#print(self.paths)
		self.ranges = 3 #hojas a leer
		self.FULLflag = False
		self.incidencias = {"bloqueo surtidor": 0,
						"atasco billetero":0,
						"impresion ticket/factura": 0,
						"parada de emergencia":0,
						"repostaje incompleto":0,
						"revelado cheque":0,
						"hoja de reclamaciones": 0,
						"predeterminado grupo":0,
						"pruebas domotica":0,
						"fallo de comunicaciones": 0,
						"problemas tecnicos": 0,
						"billete no leido. Incidencia localizada en el log": 0,
						"billete no leido. Incidencia NO localizada": 0,
						"cheque caducado": 0,
						"otra": 0
		}       
		self.resoluciones = {"apertura manual": 0,
						"toma de datos":0,
						"cheque revelado":0,
						"impreso desde cra":0,
						"responsable/expendedor informado":0,
						"pasado aviso a servicio tecnico":0,
						"estacion rearmada":0,
						"otra":0
		}
	def _resetDicts(self):
		for val,key in enumerate(self.incidencias):
			self.incidencias[key] = 0 
		for val,key in enumerate(self.resoluciones):
			self.resoluciones[key] = 0 
	def _genGroup(self):
		groups = {"total": 0}
		for i in range(24):
			if i <10:
				groups["0"+str(i)] = []
			else:
				groups[str(i)] = []
		return groups
	def _reader(self,path,groups):
		print(path)
		input()
		wb = load_workbook(filename = self.excelPATH+path)
		#print(wb.worksheets)
		if self.FULLflag == False:
			self._resetDicts()
		for i in range(self.ranges):
			ws = wb.worksheets[i]
			#print(ws)
			#print(i)
			col = ws["C"]
			#print(len(col))
			for x in range(len(col)):
				#print(str(type(col[x].value))+"||"+str(col[x].value))
				if type(col[x].value) == str:
					hora = re.search(r'\d\d:\d\d',col[x].value)
					try:
						if hora.group():
							#print(hora.group().split(":"))
							groups["total"] = groups["total"] +1
							Shora = hora.group().split(":")[0]
							groups[Shora].append(hora.group())
					except:
						'''print(str(type(col[x].value))+"||"+str(col[x].value))
						input()'''
						pass
				elif type(col[x].value) == datetime.time:
					groups["total"] = groups["total"] +1
					hora = col[x].value.hour
					if hora < 10:
						groups["0"+str(hora)].append("0"+str(hora)+":"+str(col[x].value.minute))
					else:
						groups[str(hora)].append(str(hora)+":"+str(col[x].value.minute))
			col = ws["E"]
			for x in range(len(col)):
				try:
					self.incidencias[col[x].value] = self.incidencias[col[x].value] + 1
				except KeyError:
					if col[x].value == "INCIDENCIA":
						pass
					else:
						'''print(col[x].value)
						input()'''
						self.incidencias["otra"] = self.incidencias["otra"] + 1
			col = ws["F"]
			for x in range(len(col)):
				try:
					self.resoluciones[col[x].value] = self.resoluciones[col[x].value] + 1
				except KeyError:
					if col[x].value == "RESOLUCION":
						pass
					else:
						self.resoluciones["otra"] = self.resoluciones["otra"] + 1
	def _FULLreader(self,groups):
		for FILE in self.paths:
			self._reader(FILE,groups)
	def _Percentager(self,cant, tot):
		return round((cant/tot)*100,2)
	def _xlsWriter(self,path,dic):
		wb = load_workbook(self.excelPATH+self.excelNAME)
		#print(path == "INCIDENCIAS PLENOIL.xlsx")
		if path == "INCIDENCIAS PLENOIL.xlsx":
			month = datetime.datetime.now().month
			ws = wb.worksheets[month]
		elif path == 0:
			ws = wb.worksheets[0]
		elif "MARZO" in path:
			ws = wb.worksheets[3]
		elif "ABRIL" in path:
			ws = wb.worksheets[4]
		elif "MAYO" in path:
			ws = wb.worksheets[5]
		elif "JUNIO" in path:
			ws = wb.worksheets[6]
		elif "JULIO" in path:
			ws = wb.worksheets[7]
		elif "AGOSTO" in path:
			ws = wb.worksheets[8]
		elif "SEPTIEMBRE" in path:
			ws = wb.worksheets[9]
		elif "OCTUBRE" in path:
			ws = wb.worksheets[10]
		elif "NOVIEMBRE" in path:
			ws = wb.worksheets[11]
		elif "DICIEMBRE" in path:
			ws = wb.worksheets[12]
		for i in range(24):
			if i<10:
				ws["A"+str(i+2)] = "0"+str(i)
				ws["B"+str(i+2)] = len(dic["0"+str(i)])
				cant = len(dic["0"+str(i)])
			else:
				ws["A"+str(i+2)] = str(i) 
				ws["B"+str(i+2)] = len(dic[str(i)]) 
				cant = len(dic[str(i)])
			ws["C"+str(i+2)] = self._Percentager(cant,dic["total"])
		for ind,key in enumerate(self.incidencias):
			ws["J"+str(ind+2)] = key 
			ws["K"+str(ind+2)] = self.incidencias[key]
		for ind,key in enumerate(self.resoluciones):
			ws["L"+str(ind+2)] = key 
			ws["M"+str(ind+2)] = self.resoluciones[key]

		wb.save(self.excelPATH+self.excelNAME)
	def CompleteWrite(self):
		##Escritura de cada hoja mensual
		for FILE in self.paths:
			print("Procesando: "+FILE)
			g = self._genGroup()
			self._reader(FILE,g)
			self._xlsWriter(FILE,g)
		##Escritura de la hoja de totales
		print("Procesando: TOTALES")
		self.FULLflag = True
		g = self._genGroup()
		self._FULLreader(g)
		self._xlsWriter(0,g)
	def CurrentWrite(self):
		g = self._genGroup()
		print("Procesando: INCIDENCIAS PLENOIL.xlsx")
		self._reader("INCIDENCIAS PLENOIL.xlsx", g)
		self._xlsWriter("INCIDENCIAS PLENOIL.xlsx",g)

def main():
	flux = plenFLUX()
	flux.CompleteWrite()
	
	
if __name__ == "__main__":
	main()
